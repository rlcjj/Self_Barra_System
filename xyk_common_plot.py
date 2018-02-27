# -*- coding: utf-8 -*-
"""
Created on Thu Dec 15 20:40:02 2016

@author: MouHaiMa
"""

import math
import numpy as np
import pandas as pd
from datetime import datetime
import matplotlib.pyplot as plt
from matplotlib.dates import DateFormatter, WeekdayLocator, DayLocator, MONDAY,YEARLY
import matplotlib.dates as mdate
import xyk_common_data_processing
import openpyxl

'''
------目录------
*strategy_parameter_calculate(index_list, net_value_list, date_list, strategy_name, target_name, daily = 1, simple = 0) --- 进行策略效果的综合性统计分析，并绘制策略曲线
strategy_performance(daily_date_str_list, strategy_NAV_list, benchmark_NAV_list, present_type = 0, with_plot = 1, output_format = 0, daily = 1, title_str = "Strategy Performance", \
    A_list_name = "Strategy", B_list_name = "Benchmark", annual_risk_free_rate = 0.02, Trading_points_in_one_year = 243.4) --- 策略绩效统计
cal_descriptions(daily_date_str_list, strategy_NAV_list, benchmark_NAV_list) --- 计算描述性统计结果
cal_relative_wr_and_pnl(strategy_NAV_list, benchmark_NAV_list) --- 计算相对日均胜率、盈亏比
cal_TD_and_IR(daily_date_str_list, strategy_NAV_list, benchmark_NAV_list) --- 计算跟踪偏离度TD、信息比率IR与跟踪误差TE等指标
cal_sharpe_and_sortino(daily_date_str_list, strategy_NAV_list, annual_risk_free_rate) --- 计算夏普比率、索提诺比率等指标
cal_max_drawback(daily_date_str_list, NAV_list) --- 计算最大回撤及其位置
plot_2_lines(date_str_list, A_list, B_list, daily, title_str, A_list_name, B_list_name, max_drawback_point_A = [], max_drawback_point_B = []) --- 画图：单坐标轴，双线
plot(date_str_list, A_list, B_list, daily, title_str, A_list_name, B_list_name) --- 画图
plot_time_series_plot_with_points(date_str_list, time_series_list, points_1_list, points_2_list, daily) --- 画时序图带强调点
plot_NAV_with_3_lines(date_str_list, strategy_NAV_list, benchmark_NAV_list, style_NAV_list, i) --- 绘制多线一图
cal_integrated_NAV_of_event_strategy_and_plot(dataset, date_list) --- 计算综合的事件驱动（或技术分析）类策略表现并绘图
'''

'''
***进行策略效果的综合性统计分析，并绘制策略曲线***
***注意：本函数已不再更新！！！尽量勿使用！！！***
'''
def strategy_parameter_calculate(index_list, net_value_list, date_list, strategy_name, target_name, daily = 1, simple = 0):
    '''
    ***定义常量、宏观参数***
    '''
    riskless_annual_ratio = 0.02
    
    '''
    ***数据准备与输出***
    '''
    #将benchmark的资金量调节为和我们的策略一样#
    index_adjusted_value_list = []
    start_point_of_index = index_list[0]
    for index in index_list:
        new_index = index / start_point_of_index * net_value_list[0]
        index_adjusted_value_list.append(new_index)
    
    #输出策略起始时间和终止时间，总时间长度等信息#
    start_time_str = date_list[0].strftime('%Y%m%d')
    end_time_str = date_list[-1].strftime('%Y%m%d')
    if simple == 0:
        print "*** Basic Information ***"
        print "Start Date: " + str(date_list[0].strftime('%Y-%m-%d'))
        print "End Date: " + str(date_list[-1].strftime('%Y-%m-%d'))
    Calendar_length = (datetime.strptime(end_time_str, "%Y%m%d") - datetime.strptime(start_time_str, "%Y%m%d")).days + 1
    Trading_length = len(net_value_list)
    Trading_days_in_one_year = float(Trading_length) / float(Calendar_length) * 365.0
    if simple == 0:
        print "Period Length (Calendar): " + str(Calendar_length)
        print "Period Length (Trading): " + str(Trading_length)
        print "\n"
    
    #输出策略基本情况#
    if simple == 0:
        print "*** General View ***"
        print "Start Net Value --- Strategy: " + str(net_value_list[0])
        print "End Net Value --- Strategy: " + str(net_value_list[-1])
        print "Start Net Value --- Benchmark: " + str(index_adjusted_value_list[0])
        print "End Net Value --- Benchmark: " + str(index_adjusted_value_list[-1])
        print "\n"
    
    #计算benchmark和我们的策略的总收益率与年化收益率#
    all_period_ROR_sta = net_value_list[-1] / net_value_list[0] - 1
    all_period_ROR_ben = index_adjusted_value_list[-1] / index_adjusted_value_list[0] - 1
    all_annual_ROR_sta = (all_period_ROR_sta + 1) ** (1.0 / float(Calendar_length) * 365.0) - 1
    all_annual_ROR_ben = (all_period_ROR_ben + 1) ** (1.0 / float(Calendar_length) * 365.0) - 1
    if simple == 0:
        print "*** Total Rate of Return ***"
        print "Strategy: " + str(all_period_ROR_sta)
        print "Benchmark: " + str(all_period_ROR_ben)
        print "\n"
        print "*** Annual Rate of Return ***"
        print "Strategy: " + str(all_annual_ROR_sta)
        print "Benchmark: " + str(all_annual_ROR_ben)
        print "\n"
        
    #计算benchmark和我们的策略的每日收益率#
    ROR_sta_list = []
    ROR_ben_list = []
    i = 0
    while i < len(net_value_list) - 1:
        ROR_sta = net_value_list[i + 1] / net_value_list[i] - 1
        ROR_sta_list.append(ROR_sta)
        ROR_ben = index_adjusted_value_list[i + 1] / index_adjusted_value_list[i] - 1
        ROR_ben_list.append(ROR_ben)
        i += 1
    ROR_delta_list = []
    i = 0
    while i < len(ROR_sta_list):
        ROR_delta_list.append(ROR_sta_list[i] - ROR_ben_list[i])
        i += 1
    
    #计算benchmark和我们的策略的年化波动率、夏普比率#
    ave_ROR_sta = sum(ROR_sta_list) / len(ROR_sta_list)
    ave_ROR_ben = sum(ROR_ben_list) / len(ROR_ben_list)
    ave_ROR_delta = sum(ROR_delta_list) / len(ROR_delta_list)
    total_dev_sta = 0.0
    total_dev_ben = 0.0
    total_dev_delta = 0.0
    i = 0
    while i < len(ROR_sta_list):
        total_dev_sta += (ROR_sta_list[i] - ave_ROR_sta) * (ROR_sta_list[i] - ave_ROR_sta)
        total_dev_ben += (ROR_ben_list[i] - ave_ROR_ben) * (ROR_ben_list[i] - ave_ROR_ben)
        total_dev_delta += (ROR_delta_list[i] - ave_ROR_delta) * (ROR_delta_list[i] - ave_ROR_delta)
        i += 1
    dev_sta = math.sqrt(total_dev_sta / len(ROR_sta_list))
    dev_ben = math.sqrt(total_dev_ben / len(ROR_ben_list))
    dev_delta = math.sqrt(total_dev_delta / len(ROR_delta_list))
    annual_dev_sta = dev_sta * math.sqrt(Trading_days_in_one_year)
    annual_dev_ben = dev_ben * math.sqrt(Trading_days_in_one_year)
    annual_dev_delta = dev_delta * math.sqrt(Trading_days_in_one_year)
    sharpe_sta = (all_annual_ROR_sta - riskless_annual_ratio) / annual_dev_sta
    sharpe_ben = (all_annual_ROR_ben - riskless_annual_ratio) / annual_dev_ben
    sharpe_delta = (all_annual_ROR_sta - all_annual_ROR_ben) / annual_dev_delta
    if simple == 0:
        print "*** Annual Volatility ***"
        print "Strategy: " + str(annual_dev_sta)
        print "Benchmark: " + str(annual_dev_ben)
        print "Delta: " + str(annual_dev_delta)
        print "\n"
        print "*** Annual Sharpe Ratio ***"
        print "Strategy: " + str(sharpe_sta)
        print "Benchmark: " + str(sharpe_ben)
        print "Delta: " + str(sharpe_delta)
        print "\n"
    
    #计算benchmark和我们的策略的年化下行波动率、索提诺比率#
    i = 0
    downside_ROR_sta_list = []
    downside_ROR_ben_list = []
    downside_ROR_delta_list = []
    while i < len(ROR_sta_list):
        if ROR_sta_list[i] < 0.0:
            downside_ROR_sta_list.append(ROR_sta_list[i])
        if ROR_ben_list[i] < 0.0:
            downside_ROR_ben_list.append(ROR_ben_list[i])
        if ROR_delta_list[i] < 0.0:
            downside_ROR_delta_list.append(ROR_delta_list[i])
        i += 1
    downside_ave_ROR_sta = sum(downside_ROR_sta_list) / len(downside_ROR_sta_list)
    downside_ave_ROR_ben = sum(downside_ROR_ben_list) / len(downside_ROR_ben_list)
    downside_ave_ROR_delta = sum(downside_ROR_delta_list) / len(downside_ROR_delta_list)
    downside_total_dev_sta = 0.0
    downside_total_dev_ben = 0.0
    downside_total_dev_delta = 0.0
    i = 0
    while i < len(downside_ROR_sta_list):
        downside_total_dev_sta += (downside_ROR_sta_list[i] - downside_ave_ROR_sta) * (downside_ROR_sta_list[i] - downside_ave_ROR_sta)
        i += 1
    i = 0
    while i < len(downside_ROR_ben_list):
        downside_total_dev_ben += (downside_ROR_ben_list[i] - downside_ave_ROR_ben) * (downside_ROR_ben_list[i] - downside_ave_ROR_ben)
        i += 1
    i = 0
    while i < len(downside_ROR_delta_list):
        downside_total_dev_delta += (downside_ROR_delta_list[i] - downside_ave_ROR_delta) * (downside_ROR_delta_list[i] - downside_ave_ROR_delta)
        i += 1
    downside_dev_sta = math.sqrt(downside_total_dev_sta / len(downside_ROR_sta_list))
    downside_dev_ben = math.sqrt(downside_total_dev_ben / len(downside_ROR_ben_list))
    downside_dev_delta = math.sqrt(downside_total_dev_delta / len(downside_ROR_delta_list))
    
    downside_annual_dev_sta = downside_dev_sta * math.sqrt(Trading_days_in_one_year)
    downside_annual_dev_ben = downside_dev_ben * math.sqrt(Trading_days_in_one_year)
    downside_annual_dev_delta = downside_dev_delta * math.sqrt(Trading_days_in_one_year)
    sortino_sta = (all_annual_ROR_sta - riskless_annual_ratio) / downside_annual_dev_sta
    sortino_ben = (all_annual_ROR_ben - riskless_annual_ratio) / downside_annual_dev_ben
    sortino_delta = (all_annual_ROR_sta - all_annual_ROR_ben) / downside_annual_dev_delta
    if simple == 0:
        print "*** Annual Downside Volatility ***"
        print "Strategy: " + str(downside_annual_dev_sta)
        print "Benchmark: " + str(downside_annual_dev_ben)
        print "Delta: " + str(downside_annual_dev_delta)
        print "\n"
        print "*** Annual Sortino Ratio ***"
        print "Strategy: " + str(sortino_sta)
        print "Benchmark: " + str(sortino_ben)
        print "Delta: " + str(sortino_delta)
        print "\n"
    
    #计算日均胜率和盈亏比（相对benchmark）#
    win_num_delta = 0
    lose_num_delta = 0
    win_amount_delta = 0.0
    lose_amount_delta = 0.0
    win_num_sta = 0
    lose_num_sta = 0
    win_amount_sta = 0.0
    lose_amount_sta = 0.0
    win_num_ben = 0
    lose_num_ben = 0
    win_amount_ben = 0.0
    lose_amount_ben = 0.0
    i = 0
    while i < len(ROR_sta_list):
        if ROR_sta_list[i] > ROR_ben_list[i]:
            win_num_delta += 1
            win_amount_delta += ROR_sta_list[i] - ROR_ben_list[i]
        elif ROR_sta_list[i] < ROR_ben_list[i]:
            #print i
            lose_num_delta += 1
            lose_amount_delta += ROR_ben_list[i] - ROR_sta_list[i]
        if ROR_sta_list[i] > 0:
            win_num_sta += 1
            win_amount_sta += ROR_sta_list[i] * net_value_list[i]
        elif ROR_sta_list[i] < 0:
            lose_num_sta += 1
            lose_amount_sta += ROR_sta_list[i] * net_value_list[i] * -1
        if ROR_ben_list[i] > 0:
            win_num_ben += 1
            win_amount_ben += ROR_ben_list[i] * index_adjusted_value_list[i]
        elif ROR_ben_list[i] < 0:
            lose_num_ben += 1
            lose_amount_ben += ROR_ben_list[i] * index_adjusted_value_list[i] * -1
        i += 1
    win_rate_delta = float(win_num_delta) / float(win_num_delta + lose_num_delta)
    PCR_delta = float(win_amount_delta) / float(lose_amount_delta)
    win_rate_sta = float(win_num_sta) / float(win_num_sta + lose_num_sta)
    PCR_sta = float(win_amount_sta) / float(lose_amount_sta)
    win_rate_ben = float(win_num_ben) / float(win_num_ben + lose_num_ben)
    PCR_ben = float(win_amount_ben) / float(lose_amount_ben)
    
    if simple == 0:
        print "*** Daily Win Rate ***"
        print "Strategy: " + str(win_rate_sta)
        print "Benchmark: " + str(win_rate_ben)
        print "Strategy - Benchmark: " + str(win_rate_delta)
        print "\n"
        print "*** Daily Profit and Coss Ratio ***"
        print "Strategy: " + str(PCR_sta)
        print "Benchmark: " + str(PCR_ben)
        print "Strategy - Benchmark: " + str(PCR_delta)
        print "\n"
    
    #计算最大回撤率与最大回撤位置#
    max_drawback_ratio_sta = 0.0
    max_drawback_ratio_ben = 0.0
    max_drawback_point_sta = [0, 0]
    max_drawback_point_ben = [0, 0]
    #print len(date_list)
    #print len(net_value_list)
    i = 0
    while i < len(date_list):
        j = i 
        while j < len(date_list):
            if max_drawback_ratio_sta < ((net_value_list[i] - net_value_list[j]) / net_value_list[i]):
                max_drawback_ratio_sta = ((net_value_list[i] - net_value_list[j]) / net_value_list[i])
                max_drawback_point_sta[0] = i
                max_drawback_point_sta[1] = j
            if max_drawback_ratio_ben < ((index_adjusted_value_list[i] - index_adjusted_value_list[j]) / index_adjusted_value_list[i]):
                max_drawback_ratio_ben = ((index_adjusted_value_list[i] - index_adjusted_value_list[j]) / index_adjusted_value_list[i])
                max_drawback_point_ben[0] = i
                max_drawback_point_ben[1] = j
            j += 1
        i += 1
    if simple == 0:
        print "*** Maximum Drawback Ratio ***"
        print "Strategy: " + str(max_drawback_ratio_sta) + ", between " + str(date_list[max_drawback_point_sta[0]].strftime('%Y-%m-%d')) + " and " + str(date_list[max_drawback_point_sta[1]].strftime('%Y-%m-%d'))
        print "Benchmark: " + str(max_drawback_ratio_ben) + ", between " + str(date_list[max_drawback_point_ben[0]].strftime('%Y-%m-%d')) + " and " + str(date_list[max_drawback_point_ben[1]].strftime('%Y-%m-%d'))
        print "\n"
    
    
    #计算年化收益率比最大回撤（又称收益风险比）#
    if max_drawback_ratio_sta > 0:
        ARvsMD_sta = all_annual_ROR_sta / max_drawback_ratio_sta
    else:
        ARvsMD_sta = "Infinite"
    if max_drawback_ratio_ben > 0:
        ARvsMD_ben = all_annual_ROR_ben / max_drawback_ratio_ben
    else:
        ARvsMD_ben = "Infinite"
    if simple == 0:
        print "*** Annual Return v.s. Maximum Drawback ***"
        print "Strategy: " + str(ARvsMD_sta)
        print "Benchmark: " + str(ARvsMD_ben)
        print "\n"


    '''
    ***画图***
    '''    
    #plt.rcParams['font.sas-serig']=['SimHei'] #用来正常显示中文标签
    #plt.rcParams['axes.unicode_minus']=False #用来正常显示负号 
    title_str1 = 'Net Value of ' + strategy_name + ' and ' + target_name
    fig, ax = plt.subplots(figsize=(14,3))
    ax.plot(date_list, index_adjusted_value_list, linestyle = '-', color = '#BB3D00', label = target_name, markersize = 3, linewidth = 1)
    ax.plot(date_list, net_value_list, linestyle = '-', color = '#484891', label = strategy_name, markersize=3, linewidth = 1)
    ax.plot([date_list[max_drawback_point_sta[0]], date_list[max_drawback_point_sta[1]]], [net_value_list[max_drawback_point_sta[0]], net_value_list[max_drawback_point_sta[1]]], '--d', markersize = 5, markeredgewidth = 3, linewidth = 2, color = '#A6A6D2', markeredgecolor = '#A6A6D2')
    ax.plot([date_list[max_drawback_point_ben[0]], date_list[max_drawback_point_ben[1]]], [index_adjusted_value_list[max_drawback_point_ben[0]], index_adjusted_value_list[max_drawback_point_ben[1]]], '--d', markersize = 5, markeredgewidth = 3, linewidth = 2, color = '#FF9D6F', markeredgecolor = '#FF9D6F')
    ax.xaxis.set_major_formatter(mdate.DateFormatter('%Y-%m')) #设置时间标签显示格式
    if daily == 1:
        if len(date_list) < 1000:
            plt.xticks(pd.date_range(date_list[0], date_list[-1], freq='1M'))
        elif len(date_list) < 2500:
            plt.xticks(pd.date_range(date_list[0], date_list[-1], freq='6M'))
        else:
            plt.xticks(pd.date_range(date_list[0], date_list[-1], freq='1A'))
    else:
        if len(date_list) < 40:
            plt.xticks(pd.date_range(date_list[0], date_list[-1], freq='1M'))
        elif len(date_list) < 100:
            plt.xticks(pd.date_range(date_list[0], date_list[-1], freq='6M'))
        else:
            plt.xticks(pd.date_range(date_list[0], date_list[-1], freq='1A'))

    plt.xticks(rotation = 60)
    #plt.annotate('Maximum Drawback', xy=(date_list[max_drawback_point_sta[0]], net_value_list[max_drawback_point_sta[0]]), xytext=((date_list[max_drawback_point_sta[0]] + date_list[max_drawback_point_ben[0]]) / 2, net_value_list[max_drawback_point_sta[0]] + 750), arrowprops=dict(facecolor='black', shrink=0.05),)
    #plt.annotate('Maximum Drawback', xy=(date_list[max_drawback_point_ben[0]], net_value_list[max_drawback_point_ben[0]]), xytext=((date_list[max_drawback_point_sta[0]] + date_list[max_drawback_point_ben[0]]) / 2, net_value_list[max_drawback_point_sta[0]] + 750), arrowprops=dict(facecolor='black', shrink=0.05),)
    ax.legend(loc='best')
    plt.title(title_str1)
    
    return 0

'''
***策略绩效统计***
***present_type为0时为全部数据，为1时分年度计算，为2时计算近X年***
***output_format为0时为直接输出，为1时创建excel***
'''
def strategy_performance(daily_date_str_list, strategy_NAV_list, benchmark_NAV_list, present_type = 0, with_plot = 1, output_format = 0, daily = 1, title_str = "Strategy Performance", A_list_name = "Strategy", B_list_name = "Benchmark", annual_risk_free_rate = 0.02, Trading_points_in_one_year = 243.4):
    start_index_list = []
    end_index_list = []
    if present_type == 0:
        start_index_list = [0]
        end_index_list = [len(daily_date_str_list) - 1]
    elif present_type == 1:
        start_index_list_unchanged = xyk_common_data_processing.date_partition(daily_date_str_list, 1)
        i = 0
        while i < len(start_index_list_unchanged):
            if start_index_list_unchanged[i] != 0:
                start_index_list.append(start_index_list_unchanged[i] - 1)
            else:
                start_index_list.append(start_index_list_unchanged[i])
            if i != len(start_index_list_unchanged) - 1:
                end_index_list.append(start_index_list_unchanged[i + 1] - 1)
            else:
                end_index_list.append(len(daily_date_str_list) - 1)
            i += 1
    elif present_type == 2:
        start_index_list_unchanged = xyk_common_data_processing.date_partition(daily_date_str_list, 2)
        i = 0
        while i < len(start_index_list_unchanged):
            start_index_list.append(start_index_list_unchanged[i])
            end_index_list.append(len(daily_date_str_list) - 1)
            i += 1
    if output_format == 1:
        wb = openpyxl.Workbook()
        ws = wb.active
    i = 0
    while i < len(start_index_list):
        temp_daily_date_str_list = daily_date_str_list[start_index_list[i] : end_index_list[i]] + [daily_date_str_list[end_index_list[i]]]
        temp_strategy_NAV_list = strategy_NAV_list[start_index_list[i] : end_index_list[i]] + [strategy_NAV_list[end_index_list[i]]]
        temp_benchmark_NAV_list = benchmark_NAV_list[start_index_list[i] : end_index_list[i]] + [benchmark_NAV_list[end_index_list[i]]]
        start_time_str, end_time_str, Calendar_length, Trading_length, all_period_ROR_sta, all_period_ROR_ben, all_annual_ROR_sta, all_annual_ROR_ben = cal_descriptions(temp_daily_date_str_list, temp_strategy_NAV_list, temp_benchmark_NAV_list)
        Whole_TD, Annual_TD, IR, TE = cal_TD_and_IR(temp_daily_date_str_list, temp_strategy_NAV_list, temp_benchmark_NAV_list, Trading_points_in_one_year)
        Sharpe, Sortino = cal_sharpe_and_sortino(temp_daily_date_str_list, temp_strategy_NAV_list, annual_risk_free_rate, Trading_points_in_one_year)
        win_rate_delta, pnl_delta = cal_relative_wr_and_pnl(temp_strategy_NAV_list, temp_benchmark_NAV_list)
        max_drawback_ratio_sta, max_drawback_point_sta = cal_max_drawback(temp_daily_date_str_list, temp_strategy_NAV_list)
        max_drawback_ratio_ben, max_drawback_point_ben = cal_max_drawback(temp_daily_date_str_list, temp_benchmark_NAV_list)
        if with_plot == 1:
            plot_sta_list = xyk_common_data_processing.NAV_normalize(temp_strategy_NAV_list)
            plot_ben_list = xyk_common_data_processing.NAV_normalize(temp_benchmark_NAV_list)
            plt = plot_2_lines(temp_daily_date_str_list, plot_sta_list, plot_ben_list, daily, title_str, A_list_name, B_list_name, max_drawback_point_sta, max_drawback_point_ben)
        if output_format == 0:
            print "*** Basic Information ***"
            if present_type == 0:
                print "This Period Name: " + "Whole Length"
            elif present_type == 1:
                print "This Period Name: " + "Annual Analysis --- " + end_time_str[:4]
            elif present_type == 2 and i != len(start_index_list) - 1:
                print "This Period Name: " + "Recent " + str(i + 1) + " Year"
            elif present_type == 2 and i == len(start_index_list) - 1:
                print "This Period Name: " + "From Started"
            print "Start Date: " + start_time_str
            print "End Date: " + end_time_str
            print "Total Length: " + str(Calendar_length - 1) + " (Calendar); " + str(Trading_length - 1) + " (Trading)"
            print "\n"
            print "*** Total Rate of Return ***"
            print "Strategy: " , round(all_period_ROR_sta, 4)
            print "Benchmark: ", round(all_period_ROR_ben, 4)
            print "\n"
            print "*** Annual Rate of Return ***"
            print "Strategy: ", round(all_annual_ROR_sta, 4)
            print "Benchmark: ", round(all_annual_ROR_ben, 4)
            print "\n"
            print "*** Absolute Performance ***"
            print "Sharpe Ratio: ", round(Sharpe, 3)
            print "Sortino Ratio: ", round(Sortino, 3)
            print "\n"
            print "*** Relative Performance ***"
            print "Whole Period Tracking Deviation: ", round(Whole_TD, 3)
            print "Annual Tracking Deviation: ", round(Annual_TD, 3)
            print "Annual Information Ratio: ", round(IR, 3)
            print "Annual Trackng Error: ", round(TE, 3)
            print "\n"
            print "*** Daily Relative Performance ***"
            print "Win Rate: ", round(win_rate_delta, 3)
            print "Profit & Loss: ", round(pnl_delta, 3)
            print "\n"
            print "*** Drawback Indexes ***"
            print "Max Drawback Ratio of Strategy: ", round(max_drawback_ratio_sta, 4), " , from ", daily_date_str_list[max_drawback_point_sta[0]], " to ", daily_date_str_list[max_drawback_point_sta[1]]
            print "Max Drawback Ratio of Benchmark: ", round(max_drawback_ratio_ben, 4), " , from ", daily_date_str_list[max_drawback_point_ben[0]], " to ", daily_date_str_list[max_drawback_point_ben[1]]
            print "\n"
            print "------------------------------------"
            print "\n"
        elif output_format == 1:
            if i == 0:
                ws.cell(row = 1, column = 1, value = "This Period Name")
                ws.cell(row = 2, column = 1, value = u"时间段")
                ws.cell(row = 1, column = 2, value = "Start Date")
                ws.cell(row = 2, column = 2, value = u"开始时间")
                ws.cell(row = 1, column = 3, value = "End Date")
                ws.cell(row = 2, column = 3, value = u"截止时间")
                ws.cell(row = 1, column = 4, value = "Total Length (Calendar)")
                ws.cell(row = 2, column = 4, value = u"日历日总数量")
                ws.cell(row = 1, column = 5, value = "Total Length (Trading)")
                ws.cell(row = 2, column = 5, value = u"交易日总数量")
                ws.cell(row = 1, column = 6, value = "Total Rate of Return - Strategy")
                ws.cell(row = 2, column = 6, value = u"策略总收益率")
                ws.cell(row = 1, column = 7, value = "Total Rate of Return - Benchmark")
                ws.cell(row = 2, column = 7, value = u"基准总收益率")
                ws.cell(row = 1, column = 8, value = "Annual Rate of Return - Strategy")
                ws.cell(row = 2, column = 8, value = u"策略年化收益率")
                ws.cell(row = 1, column = 9, value = "Annual Rate of Return - Benchmark")
                ws.cell(row = 2, column = 9, value = u"基准年化收益率")
                ws.cell(row = 1, column = 10, value = "Sharpe Ratio")
                ws.cell(row = 2, column = 10, value = u"夏普比率")
                ws.cell(row = 1, column = 11, value = "Sortino Ratio")
                ws.cell(row = 2, column = 11, value = u"索提诺比率")
                ws.cell(row = 1, column = 12, value = "Whole Period tracking deviation")
                ws.cell(row = 2, column = 12, value = u"总跟踪偏离度")
                ws.cell(row = 1, column = 13, value = "Annual Tracking Deviation")
                ws.cell(row = 2, column = 13, value = u"年化跟踪偏离度")
                ws.cell(row = 1, column = 14, value = "Annual Information Ratio")
                ws.cell(row = 2, column = 14, value = u"年化信息比率")
                ws.cell(row = 1, column = 15, value = "Annual Trackng Error")
                ws.cell(row = 2, column = 15, value = u"年化跟踪误差")
                ws.cell(row = 1, column = 16, value = "Win Rate")
                ws.cell(row = 2, column = 16, value = u"日频胜率")
                ws.cell(row = 1, column = 17, value = "Profit & Loss")
                ws.cell(row = 2, column = 17, value = u"日频盈亏比")
                ws.cell(row = 1, column = 18, value = "Max Drawback Ratio - Strategy")
                ws.cell(row = 2, column = 18, value = u"策略最大回撤")
                ws.cell(row = 1, column = 19, value = "Max Drawback Period - Strategy")
                ws.cell(row = 2, column = 19, value = u"策略最大回撤发生时间")
                ws.cell(row = 1, column = 20, value = "Max Drawback Ratio - Benchmark")
                ws.cell(row = 2, column = 20, value = u"基准最大回撤")
                ws.cell(row = 1, column = 21, value = "Max Drawback Period - Benchmark")
                ws.cell(row = 2, column = 21, value = u"基准最大回撤发生时间")
            
            if present_type == 0:
                ws.cell(row = i + 3, column = 1, value = "Whole Length")
            elif present_type == 1:
                ws.cell(row = i + 3, column = 1, value = end_time_str[:4])
            elif present_type == 2 and i != len(start_index_list) - 1:
                ws.cell(row = i + 3, column = 1, value = "Recent " + str(i + 1) + " Year")
            elif present_type == 2 and i == len(start_index_list) - 1:
                ws.cell(row = i + 3, column = 1, value = "From Started")
            ws.cell(row = i + 3, column = 2, value = start_time_str)
            ws.cell(row = i + 3, column = 3, value = end_time_str)
            ws.cell(row = i + 3, column = 4, value = Calendar_length - 1)
            ws.cell(row = i + 3, column = 5, value = Trading_length - 1)
            ws.cell(row = i + 3, column = 6, value = round(all_period_ROR_sta, 4))
            ws.cell(row = i + 3, column = 7, value = round(all_period_ROR_ben, 4))
            ws.cell(row = i + 3, column = 8, value = round(all_annual_ROR_sta, 4))
            ws.cell(row = i + 3, column = 9, value = round(all_annual_ROR_ben, 4))
            ws.cell(row = i + 3, column = 10, value = round(Sharpe, 3))
            ws.cell(row = i + 3, column = 11, value = round(Sortino, 3))
            ws.cell(row = i + 3, column = 12, value = round(Whole_TD, 3))
            ws.cell(row = i + 3, column = 13, value = round(Annual_TD, 3))
            ws.cell(row = i + 3, column = 14, value = round(IR, 3))
            ws.cell(row = i + 3, column = 15, value = round(TE, 3))
            ws.cell(row = i + 3, column = 16, value = round(win_rate_delta, 3))
            ws.cell(row = i + 3, column = 17, value = round(pnl_delta, 3))
            ws.cell(row = i + 3, column = 18, value = round(max_drawback_ratio_sta, 4))
            ws.cell(row = i + 3, column = 19, value = "from " + str(daily_date_str_list[max_drawback_point_sta[0]]) + " to " + str(daily_date_str_list[max_drawback_point_sta[1]]))
            ws.cell(row = i + 3, column = 20, value = round(max_drawback_ratio_ben, 4))
            ws.cell(row = i + 3, column = 21, value = "from " + str(daily_date_str_list[max_drawback_point_ben[0]]) + " to " + str(daily_date_str_list[max_drawback_point_ben[1]]))
            
            if with_plot == 1:
                plt.savefig("No_" + str(i) + "_performance.png", dpi = 60)
                img = openpyxl.drawing.image.Image("No_" + str(i) + "_performance.png")
                img.anchor(ws.cell(row = 10 * i + len(start_index_list) + 6, column = 1))
                ws.add_image(img)
                
            wb.save('strategy_performance.xlsx')
        i += 1
    return 0

'''
***计算描述性统计结果***
'''
def cal_descriptions(daily_date_str_list, strategy_NAV_list, benchmark_NAV_list):
    normalized_strategy_NAV_list = xyk_common_data_processing.NAV_normalize(strategy_NAV_list)
    normalized_benchmark_NAV_list = xyk_common_data_processing.NAV_normalize(benchmark_NAV_list)
    start_time_str = daily_date_str_list[0]
    end_time_str = daily_date_str_list[-1]
    Calendar_length = (datetime.strptime(end_time_str, "%Y%m%d") - datetime.strptime(start_time_str, "%Y%m%d")).days + 1
    Trading_length = len(daily_date_str_list)
    all_period_ROR_sta = normalized_strategy_NAV_list[-1] / normalized_strategy_NAV_list[0] - 1
    all_period_ROR_ben = normalized_benchmark_NAV_list[-1] / normalized_benchmark_NAV_list[0] - 1
    all_annual_ROR_sta = (all_period_ROR_sta + 1) ** (1.0 / float(Calendar_length) * 365.0) - 1
    all_annual_ROR_ben = (all_period_ROR_ben + 1) ** (1.0 / float(Calendar_length) * 365.0) - 1
    return start_time_str, end_time_str, Calendar_length, Trading_length, all_period_ROR_sta, all_period_ROR_ben, all_annual_ROR_sta, all_annual_ROR_ben

'''
***计算相对日均胜率、盈亏比***
'''
def cal_relative_wr_and_pnl(strategy_NAV_list, benchmark_NAV_list):
    win_num_delta = 0
    lose_num_delta = 0
    win_amount_delta = 0.0
    lose_amount_delta = 0.0
    i = 1
    while i < len(strategy_NAV_list):
        ROR_sta = strategy_NAV_list[i] / strategy_NAV_list[i - 1] - 1.0
        ROR_ben = benchmark_NAV_list[i] / benchmark_NAV_list[i - 1] - 1.0
        if ROR_sta > ROR_ben + 0.00001:
            win_num_delta += 1
            win_amount_delta += ROR_sta - ROR_ben
        elif ROR_sta < ROR_ben - 0.00001:
            lose_num_delta += 1
            lose_amount_delta += ROR_ben - ROR_sta
        else:
            pass
        i += 1
         
    win_rate_delta = float(win_num_delta) / float(win_num_delta + lose_num_delta)
    if win_num_delta == 0:
        pnl_delta = 0.0
    elif lose_num_delta == 0:
        pnl_delta = 10000.0
    else:
        pnl_delta = (float(win_amount_delta) / float(win_num_delta)) / (float(lose_amount_delta) / float(lose_num_delta))
    return win_rate_delta, pnl_delta

'''
***计算跟踪偏离度TD、信息比率IR与跟踪误差TE等指标***
'''
def cal_TD_and_IR(daily_date_str_list, strategy_NAV_list, benchmark_NAV_list, Trading_points_in_one_year = 243.4):
    normalized_strategy_NAV_list = xyk_common_data_processing.NAV_normalize(strategy_NAV_list)
    normalized_benchmark_NAV_list = xyk_common_data_processing.NAV_normalize(benchmark_NAV_list)
    start_time_str = daily_date_str_list[0]
    end_time_str = daily_date_str_list[-1]
    Calendar_length = (datetime.strptime(end_time_str, "%Y%m%d") - datetime.strptime(start_time_str, "%Y%m%d")).days + 1
    Trading_length = len(daily_date_str_list)
    if Calendar_length < 360:
        Trading_days_in_one_year = Trading_points_in_one_year
    else:
        Trading_days_in_one_year = float(Trading_length) / float(Calendar_length) * 365.0
    
    #计算benchmark和我们的策略的总收益率与年化收益率#
    all_period_ROR_sta = normalized_strategy_NAV_list[-1] / normalized_strategy_NAV_list[0] - 1.0
    all_period_ROR_ben = normalized_benchmark_NAV_list[-1] / normalized_benchmark_NAV_list[0] - 1.0
    all_period_ROR_delta = all_period_ROR_sta - all_period_ROR_ben
    all_annual_ROR_delta = (all_period_ROR_delta + 1.0) ** (1.0 / float(Calendar_length) * 365.0) - 1.0
    
    #计算benchmark和我们的策略的年化波动率、信息比率#
    ROR_list_sta = xyk_common_data_processing.from_NAV_to_ROR(normalized_strategy_NAV_list)
    ROR_list_ben = xyk_common_data_processing.from_NAV_to_ROR(normalized_benchmark_NAV_list)
    ROR_list_delta = xyk_common_data_processing.plus_and_minus_between_list(ROR_list_sta, ROR_list_ben, "minus")
    if len(ROR_list_delta) <= 1:
        IR_delta = 10000.0
        annual_dev_delta = 0.0
    else:
        ROR_dev_delta = xyk_common_data_processing.dev_n_1(ROR_list_delta)
        annual_dev_delta = ROR_dev_delta * math.sqrt(Trading_days_in_one_year)
        IR_delta = all_annual_ROR_delta / annual_dev_delta
    
    return all_period_ROR_delta, all_annual_ROR_delta, IR_delta, annual_dev_delta

'''
***计算夏普比率、索提诺比率等指标***
'''
def cal_sharpe_and_sortino(daily_date_str_list, strategy_NAV_list, annual_risk_free_rate = 0.02, Trading_points_in_one_year = 243.4):
    normalized_strategy_NAV_list = xyk_common_data_processing.NAV_normalize(strategy_NAV_list)
    start_time_str = daily_date_str_list[0]
    end_time_str = daily_date_str_list[-1]
    Calendar_length = (datetime.strptime(end_time_str, "%Y%m%d") - datetime.strptime(start_time_str, "%Y%m%d")).days + 1
    Trading_length = len(daily_date_str_list)
    if Calendar_length < 360:
        Trading_days_in_one_year = Trading_points_in_one_year
    else:
        Trading_days_in_one_year = float(Trading_length) / float(Calendar_length) * 365.0
    
    #计算benchmark和我们的策略的总收益率与年化收益率#
    all_period_ROR_sta = normalized_strategy_NAV_list[-1] / normalized_strategy_NAV_list[0] - 1.0
    all_annual_ROR_sta = (all_period_ROR_sta + 1.0) ** (1.0 / float(Calendar_length) * 365.0) - 1.0
    
    #计算benchmark和我们的策略的年化波动率、夏普比率#
    ROR_list_sta = xyk_common_data_processing.from_NAV_to_ROR(normalized_strategy_NAV_list)
    if len(ROR_list_sta) <= 1:
        Sharpe_sta = 10000.0
    else:
        ROR_dev_sta = xyk_common_data_processing.dev_n_1(ROR_list_sta)
        annual_dev_sta = ROR_dev_sta * math.sqrt(Trading_days_in_one_year)
        Sharpe_sta = (all_annual_ROR_sta - annual_risk_free_rate) / annual_dev_sta
    
    #计算benchmark和我们的策略的年化下行波动率、索提诺比率#
    downside_ROR_list_sta = []
    i = 0
    while i < len(ROR_list_sta):
        if ROR_list_sta[i] < annual_risk_free_rate / Trading_days_in_one_year:
            downside_ROR_list_sta.append(ROR_list_sta[i] - annual_risk_free_rate / Trading_days_in_one_year)
        i += 1
    if len(downside_ROR_list_sta) <= 1:
        Sortino_sta = 10000.0
    else:
        downside_ROR_dev_sta = xyk_common_data_processing.dev_n_1(downside_ROR_list_sta)
        downside_annual_dev_sta = downside_ROR_dev_sta * math.sqrt(Trading_days_in_one_year)
        Sortino_sta = (all_annual_ROR_sta - annual_risk_free_rate) / downside_annual_dev_sta
    
    return Sharpe_sta, Sortino_sta
            
'''
***计算最大回撤及其位置***
'''
def cal_max_drawback(daily_date_str_list, NAV_list):
    max_drawback_ratio = 0.0
    max_drawback_point = [0, 0]
    i = 0
    while i < len(daily_date_str_list):
        j = i 
        while j < len(daily_date_str_list):
            if max_drawback_ratio < ((NAV_list[i] - NAV_list[j]) / NAV_list[i]):
                max_drawback_ratio = ((NAV_list[i] - NAV_list[j]) / NAV_list[i])
                max_drawback_point[0] = i
                max_drawback_point[1] = j
            j += 1
        i += 1
    return max_drawback_ratio, max_drawback_point

'''
***画图：单坐标轴，双线***
'''  
def plot_2_lines(date_str_list, A_list, B_list, daily, title_str, A_list_name, B_list_name, max_drawback_point_A = [], max_drawback_point_B = []):
    date_list = xyk_common_data_processing.change_date_format("str", "datetime", date_str_list)
    fig, ax = plt.subplots(figsize=(14,3))
    ax.plot(date_list, A_list, linestyle = '-', color = '#BB3D00', label = A_list_name, markersize = 3, linewidth = 1)
    ax.plot(date_list, B_list, linestyle = '-', color = '#484891', label = B_list_name, markersize=3, linewidth = 1)
    if len(max_drawback_point_A) != 0:
        ax.plot([date_list[max_drawback_point_A[0]], date_list[max_drawback_point_A[1]]], [A_list[max_drawback_point_A[0]], A_list[max_drawback_point_A[1]]], '--d', markersize = 5, markeredgewidth = 3, linewidth = 2, color = '#FF9D6F', markeredgecolor = '#FF9D6F')
    if len(max_drawback_point_B) != 0:    
        ax.plot([date_list[max_drawback_point_B[0]], date_list[max_drawback_point_B[1]]], [B_list[max_drawback_point_B[0]], B_list[max_drawback_point_B[1]]], '--d', markersize = 5, markeredgewidth = 3, linewidth = 2, color = '#A6A6D2', markeredgecolor = '#A6A6D2')
    ax.xaxis.set_major_formatter(mdate.DateFormatter('%Y-%m')) #设置时间标签显示格式
    if daily == 1:
        if len(date_list) < 1000:
            plt.xticks(pd.date_range(date_list[0], date_list[-1], freq='1M').shift(1, freq='D'))
        elif len(date_list) < 2500:
            plt.xticks(pd.date_range(date_list[0], date_list[-1], freq='6M').shift(1, freq='D'))
        else:
            plt.xticks(pd.date_range(date_list[0], date_list[-1], freq='1A').shift(1, freq='D'))
    else:
        if len(date_list) < 40:
            plt.xticks(pd.date_range(date_list[0], date_list[-1], freq='1M').shift(1, freq='D'))
        elif len(date_list) < 100:
            plt.xticks(pd.date_range(date_list[0], date_list[-1], freq='6M').shift(1, freq='D'))
        else:
            plt.xticks(pd.date_range(date_list[0], date_list[-1], freq='1A').shift(1, freq='D'))

    #plt.xticks(rotation = 60)
    
    #plt.annotate('Maximum Drawback', xy=(date_list[max_drawback_point_sta[0]], net_value_list[max_drawback_point_sta[0]]), xytext=((date_list[max_drawback_point_sta[0]] + date_list[max_drawback_point_ben[0]]) / 2, net_value_list[max_drawback_point_sta[0]] + 750), arrowprops=dict(facecolor='black', shrink=0.05),)
    #plt.annotate('Maximum Drawback', xy=(date_list[max_drawback_point_ben[0]], net_value_list[max_drawback_point_ben[0]]), xytext=((date_list[max_drawback_point_sta[0]] + date_list[max_drawback_point_ben[0]]) / 2, net_value_list[max_drawback_point_sta[0]] + 750), arrowprops=dict(facecolor='black', shrink=0.05),)
    ax.legend(loc='best')
    plt.title(title_str)
    return plt

'''
***画图：双坐标轴，左轴为红，右轴为蓝***
'''   
def plot(date_str_list, A_list, B_list, daily, title_str, A_list_name, B_list_name):
    datetime_list = xyk_common_data_processing.change_date_format("str", "datetime", date_str_list)
    fig, ax1 = plt.subplots(figsize=(14,3))
    ax1.plot(datetime_list, A_list, linestyle = '-', color = '#BB3D00', label = A_list_name, markersize = 3, linewidth = 1)
    ax1.set_ylabel(A_list_name)
    ax2 = ax1.twinx()
    ax2.plot(datetime_list, B_list, 'r', linestyle = '-', color = '#484891', label = B_list_name, markersize = 3, linewidth = 1)
    ax2.set_ylabel(B_list_name)
    ax1.xaxis.set_major_formatter(mdate.DateFormatter('%Y-%m')) #设置时间标签显示格式
    if daily == 1:
        if len(datetime_list) < 1000:
            plt.xticks(pd.date_range(datetime_list[0], datetime_list[-1], freq='1M').shift(1, freq='D'))
        elif len(datetime_list) < 2500:
            plt.xticks(pd.date_range(datetime_list[0], datetime_list[-1], freq='6M').shift(1, freq='D'))
        else:
            plt.xticks(pd.date_range(datetime_list[0], datetime_list[-1], freq='1A').shift(1, freq='D'))
    else:
        if len(datetime_list) < 40:
            plt.xticks(pd.date_range(datetime_list[0], datetime_list[-1], freq='1M').shift(1, freq='D'))
        elif len(datetime_list) < 100:
            plt.xticks(pd.date_range(datetime_list[0], datetime_list[-1], freq='6M').shift(1, freq='D'))
        else:
            plt.xticks(pd.date_range(datetime_list[0], datetime_list[-1], freq='1A').shift(1, freq='D'))
    
    ax1.legend(loc=2)
    ax2.legend(loc=1)
    plt.title(title_str)
    return 0

'''
***绘制时间序列图与对应的强调点***
'''
def plot_time_series_plot_with_points(date_str_list, time_series_list, points_1_list, points_2_list, daily):
    date_dt_list = xyk_common_data_processing.change_date_format("str", "datetime", date_str_list)
    
    title_str = 'Time Series'
    fig, ax = plt.subplots(figsize=(14,3))
    ax.plot(date_dt_list, time_series_list, linestyle = '-', color = '#484891', label = 'time series', markersize = 1, linewidth = 1)
    if len(points_1_list) != 0:
        points_1_dt_list = []
        points_1_ts_list = []
        i = 0
        while i < len(points_1_list):
            points_1_dt_list.append(date_dt_list[points_1_list[i]])
            points_1_ts_list.append(time_series_list[points_1_list[i]])
            i += 1
        ax.plot(points_1_dt_list, points_1_ts_list, 'rx', label = 'max_points', markersize = 10)
    if len(points_2_list) != 0:
        points_2_dt_list = []
        points_2_ts_list = []
        i = 0
        while i < len(points_2_list):
            points_2_dt_list.append(date_dt_list[points_2_list[i]])
            points_2_ts_list.append(time_series_list[points_2_list[i]])
            i += 1
        ax.plot(points_2_dt_list, points_2_ts_list, 'gx', label = 'min_points', markersize = 10)
    ax.xaxis.set_major_formatter(mdate.DateFormatter('%Y-%m')) #设置时间标签显示格式
    if daily == 1:
        if len(date_dt_list) < 1000:
            plt.xticks(pd.date_range(date_dt_list[0], date_dt_list[-1], freq='1M').shift(1, freq='D'))
        elif len(date_dt_list) < 2500:
            plt.xticks(pd.date_range(date_dt_list[0], date_dt_list[-1], freq='6M').shift(1, freq='D'))
        else:
            plt.xticks(pd.date_range(date_dt_list[0], date_dt_list[-1], freq='1A').shift(1, freq='D'))
    else:
        if len(date_dt_list) < 40:
            plt.xticks(pd.date_range(date_dt_list[0], date_dt_list[-1], freq='1M').shift(1, freq='D'))
        elif len(date_dt_list) < 100:
            plt.xticks(pd.date_range(date_dt_list[0], date_dt_list[-1], freq='6M').shift(1, freq='D'))
        else:
            plt.xticks(pd.date_range(date_dt_list[0], date_dt_list[-1], freq='1A').shift(1, freq='D'))
    ax.legend(loc='best')
    plt.title(title_str)
    
    return 0

'''
***绘制多线一图***
'''
def plot_NAV_with_3_lines(date_str_list, strategy_NAV_list, benchmark_NAV_list, style_NAV_list, i):
    normalized_strategy_NAV_list = xyk_common_data_processing.NAV_normalize(strategy_NAV_list)
    normalized_benchmark_NAV_list = xyk_common_data_processing.NAV_normalize(benchmark_NAV_list)
    normalized_style_NAV_list = xyk_common_data_processing.NAV_normalize(style_NAV_list)
    date_dt_list = xyk_common_data_processing.change_date_format("str", "datetime", date_str_list)
    
    title_str1 = 'NAV of Style ' + str(i + 1)
    fig, ax = plt.subplots(figsize=(14,3))
    ax.plot(date_dt_list, normalized_strategy_NAV_list, linestyle = '-', color = '#BB3D00', label = 'dynamic strategy', markersize = 1, linewidth = 1)
    ax.plot(date_dt_list, normalized_benchmark_NAV_list, linestyle = '-', color = '#484891', label = 'momentum strategy', markersize = 1, linewidth = 1)
    ax.plot(date_dt_list, normalized_style_NAV_list, linestyle = '-', color = '#FFD700', label = 'holding strategy', markersize = 1, linewidth = 1)
    ax.xaxis.set_major_formatter(mdate.DateFormatter('%Y-%m')) #设置时间标签显示格式
    if len(date_dt_list) < 1000:
        plt.xticks(pd.date_range(date_dt_list[0], date_dt_list[-1], freq='1M').shift(1, freq='D'))
    elif len(date_dt_list) < 2500:
        plt.xticks(pd.date_range(date_dt_list[0], date_dt_list[-1], freq='6M').shift(1, freq='D'))
    else:
        plt.xticks(pd.date_range(date_dt_list[0], date_dt_list[-1], freq='1A').shift(1, freq='D'))

    plt.xticks(rotation = 60)
    ax.legend(loc='best')
    plt.title(title_str1)
    
    return 0

'''
***计算综合的事件驱动（或技术分析）类策略表现并绘图***
***dataset的形式为[开始时间，结束时间，收益]，要求开始结束时间均在date_list中，但date_list可不为日线***
'''
def cal_integrated_NAV_of_event_strategy_and_plot(dataset, date_list):
    every_unit_ROR_list = [[] for j in range(len(date_list) - 1)]
    
    for trading in dataset:
        start_index = date_list.index(trading[0])
        end_index = date_list.index(trading[1])
        R0 = xyk_common_data_processing.cal_R0_from_TR(trading[2], end_index - start_index)
        i = start_index
        while i < end_index:
            every_unit_ROR_list[i].append(R0)
            i += 1
            
    ROR_list = []
    Amount_list = []
    for unit in every_unit_ROR_list:
        if len(unit) == 0:
            ROR_list.append(0.0)
            Amount_list.append(0)
        else:
            ROR_list.append(np.mean(unit))
            Amount_list.append(len(unit))
            
    NAV_list = xyk_common_data_processing.from_ROR_to_NAV(ROR_list)
    weekly_dt = xyk_common_data_processing.change_date_format("str", "datetime", date_list[:-1])
    plot(weekly_dt, NAV_list[1:], Amount_list, 0, "NAV and Amount", "NAV", "Amount")
    
    return 0